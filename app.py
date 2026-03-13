"""
MTC Validator — app.py (v0.3 Production)
Complete system with authentication, PDF OCR, history tracking, and professional reports.
"""

import streamlit as st
import pandas as pd
import io
import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth
from normas import NORMAS
from validator import validate_mtc, clean_dataframe, detect_norma
from report import generate_pdf
from database import init_db, save_validacion, get_historial, get_stats, delete_validacion, delete_all_usuario
from ocr_pdf import extract_from_pdf

# ── Config ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MTC Validator v0.3",
    page_icon="⬡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── AUTHENTICATION ───────────────────────────────────────────────────────────
@st.cache_resource
def load_authenticator():
    try:
        with open("config.yaml") as config_file:
            config = yaml.load(config_file, Loader=SafeLoader)
            if not config or 'credentials' not in config:
                raise ValueError("Invalid YAML structure: missing 'credentials' key")
    except FileNotFoundError:
        st.error("❌ config.yaml not found. Create it in the project root directory.")
        st.stop()
    except Exception as e:
        st.error(f"❌ Error loading config.yaml: {str(e)}")
        st.stop()
    
    try:
        # Ensure required keys
        if 'cookie' not in config:
            config['cookie'] = {'expiry_days': 30, 'key': 'mtc_secret_key_2024', 'name': 'mtc_auth'}
        if 'pre-authorized' not in config:
            config['pre-authorized'] = []
        
        authenticator = stauth.Authenticate(
            config['credentials'],
            config['cookie']['name'],
            config['cookie']['key'],
            config['cookie']['expiry_days'],
            config['pre-authorized']
        )
        return authenticator
    except Exception as e:
        st.error(f"❌ Authentication setup error: {str(e)}")
        st.stop()

authenticator = load_authenticator()
name, authentication_status, username = authenticator.login("main", "sidebar")

if authentication_status is False:
    st.error("❌ Usuario o contraseña incorrectos. Intenta de nuevo.")
    st.stop()
elif authentication_status is None:
    st.warning("👤 Ingresa tu usuario y contraseña para continuar.")
    st.stop()

# Load user email from config after authentication
if authentication_status and username:
    with open("config.yaml") as config_file:
        config = yaml.load(config_file, Loader=SafeLoader)
        user_data = config['credentials']['usernames'].get(username, {})
        st.session_state.email = user_data.get('email', 'no-email@mtcvalidator.com')
        st.session_state.username = username
        st.session_state.name = name

# Initialize database on first load
if "db_initialized" not in st.session_state:
    init_db()
    st.session_state.db_initialized = True

# ── CSS ──────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;500;600&family=Syne:wght@400;600;700;800&display=swap');
:root {
  --accent: #00d4a0; --red: #ff4d6a; --yellow: #f5a623;
  --surface: #0d1218; --border: #1e2835;
  --mono: 'JetBrains Mono', monospace; --display: 'Syne', sans-serif;
}
html, body, [class*="css"] { font-family: var(--mono) !important; }
.main .block-container { padding: 1.5rem 2rem 2rem !important; max-width: 100% !important; }
header[data-testid="stHeader"] { background: #0d1218 !important; border-bottom: 1px solid #1e2835; }
section[data-testid="stSidebar"] { background: #0d1218 !important; border-right: 1px solid #1e2835 !important; }
div[data-baseweb="select"] > div {
  background: #080c10 !important; border: 1px solid #263040 !important;
  border-radius: 4px !important; font-family: var(--mono) !important; color: #00d4a0 !important;
}
[data-testid="metric-container"] {
  background: #111820 !important; border: 1px solid #1e2835 !important;
  border-radius: 6px !important; padding: 0.75rem 1rem !important;
}
[data-testid="metric-container"] label { font-size: 0.65rem !important; letter-spacing: 0.1em !important; color: #4a6070 !important; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #00d4a0 !important; font-size: 1.4rem !important; font-weight: 600 !important; }
.stDownloadButton > button, .stButton > button {
  background: #00d4a0 !important; color: #080c10 !important; border: none !important;
  border-radius: 4px !important; font-family: var(--mono) !important;
  font-weight: 600 !important; font-size: 0.8rem !important; letter-spacing: 0.06em !important;
}
[data-testid="stDataFrame"] { border: 1px solid #1e2835 !important; border-radius: 6px !important; }
hr { border-color: #1e2835 !important; }
::-webkit-scrollbar { width: 4px; } ::-webkit-scrollbar-thumb { background: #263040; border-radius: 2px; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
def verdict_banner(texto, subtexto, tipo="pass"):
    cfg = {
        "pass":    ("#00d4a0", "rgba(0,212,160,0.07)",  "rgba(0,212,160,0.3)"),
        "fail":    ("#ff4d6a", "rgba(255,77,106,0.10)", "rgba(255,77,106,0.3)"),
        "warning": ("#f5a623", "rgba(245,166,35,0.08)", "rgba(245,166,35,0.3)"),
    }
    color, bg, border = cfg.get(tipo, cfg["pass"])
    return f"""
    <div style="background:{bg};border:2px solid {border};border-radius:6px;
                padding:14px 18px;margin-bottom:1rem;">
      <div style="font-family:'Syne',sans-serif;font-size:0.95rem;font-weight:700;color:{color};">
        {texto}
      </div>
      <div style="font-size:0.7rem;color:#6a8090;margin-top:2px;">{subtexto}</div>
    </div>"""

def generate_template(norma_key: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "MTC"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30

    for col, val in [("A1", "elemento"), ("B1", "valor"), ("C1", "referencia")]:
        ws[col] = val
        ws[col].font = Font(bold=True, color="00D4A0", name="Courier New")
        ws[col].fill = PatternFill("solid", start_color="0D1218")
        ws[col].alignment = Alignment(horizontal="center")

    norma = NORMAS[norma_key]
    for i, (elem, limits) in enumerate(norma.items(), start=2):
        ws[f"A{i}"] = elem
        ws[f"B{i}"] = ""
        ws[f"C{i}"] = f"Min: {limits['min']} | Max: {limits['max']}"
        ws[f"A{i}"].font = Font(name="Courier New", size=10)
        ws[f"B{i}"].font = Font(name="Courier New", size=10, bold=True)
        ws[f"C{i}"].font = Font(name="Courier New", size=9, color="4A6070")
        ws[f"B{i}"].alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="margin-bottom:1.5rem;">
      <div style="font-size:0.6rem;color:#4a6070;letter-spacing:0.12em;margin-bottom:4px;">&gt; SISTEMA ACTIVO</div>
      <div style="font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:800;color:#00d4a0;line-height:1;letter-spacing:-0.02em;">MTC<br>Validator</div>
      <div style="font-size:0.6rem;color:#4a6070;margin-top:4px;letter-spacing:0.06em;">v0.3 — ASTM / SAE / NMX</div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown(f'<div style="font-size:0.65rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:8px;">👤 {st.session_state.name}</div>', unsafe_allow_html=True)
    
    if st.button("🚪 Logout", use_container_width=True):
        authenticator.logout("Logout", "sidebar")
        st.rerun()

    st.markdown("<hr>", unsafe_allow_html=True)
    
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("VERSIÓN", "0.3")
    with col_b:
        st.metric("NORMAS", len(NORMAS))

# ── MAIN TABS ────────────────────────────────────────────────────────────────
tab_validador, tab_historial, tab_config = st.tabs(["✓ Validador", "📊 Historial", "⚙️ Configuración"])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1: VALIDADOR
# ═══════════════════════════════════════════════════════════════════════════════
with tab_validador:
    st.markdown("""
    <div style="margin-bottom:1.5rem;">
      <div style="font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:700;color:#c8d8e8;">
        Validador de Certificados MTC
      </div>
      <div style="font-size:0.7rem;color:#4a6070;margin-top:2px;">
        Soporta Excel (.xlsx) y PDFs digitales/escaneados
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="font-size:0.6rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:6px;">SELECCIONA NORMA</div>', unsafe_allow_html=True)
    norma_seleccionada = st.selectbox(
        "Norma", options=list(NORMAS.keys()), index=0, label_visibility="collapsed", key="norma_sel"
    )

    st.markdown('<div style="font-size:0.6rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:6px;margin-top:12px;">TOLERANCIAS REFERENCIA</div>', unsafe_allow_html=True)
    ref_data = [{"Elemento": k, "Mín": v["min"], "Máx": v["max"]} for k, v in NORMAS[norma_seleccionada].items()]
    st.dataframe(pd.DataFrame(ref_data), hide_index=True, use_container_width=True, height=200)

    st.markdown('<div style="font-size:0.6rem;color:#4a6070;letter-spacing:0.1em;margin-top:12px;margin-bottom:6px;">DESCARGAR PLANTILLA</div>', unsafe_allow_html=True)
    template_bytes = generate_template(norma_seleccionada)
    st.download_button(
        label="↓ Plantilla Excel",
        data=template_bytes,
        file_name=f"plantilla_mtc_{norma_seleccionada}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.markdown("<hr>", unsafe_allow_html=True)

    st.markdown('<div style="font-size:0.65rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:8px;">CARGAR CERTIFICADO</div>', unsafe_allow_html=True)
    
    tab_excel, tab_pdf = st.tabs(["📄 Excel", "📕 PDF"])
    
    archivo_a_validar = None
    metodo_ingreso = None
    
    with tab_excel:
        archivo_excel = st.file_uploader(
            "Excel", type=["xlsx", "xls"], label_visibility="collapsed", key="excel_upload"
        )
        archivo_a_validar = archivo_excel
        metodo_ingreso = "excel" if archivo_excel else None
    
    with tab_pdf:
        archivo_pdf = st.file_uploader(
            "PDF", type=["pdf"], label_visibility="collapsed", key="pdf_upload"
        )
        archivo_a_validar = archivo_pdf
        metodo_ingreso = "pdf" if archivo_pdf else None

    st.markdown("<hr>", unsafe_allow_html=True)

    if archivo_a_validar is not None:
        try:
            if metodo_ingreso == "pdf":
                st.info("📥 Extrayendo datos del PDF...")
                df_clean, metodo_ocr = extract_from_pdf(archivo_a_validar.read())
                
                if df_clean.empty:
                    st.error("❌ No se extrajeron datos del PDF.")
                    st.stop()
                
                st.success(f"✓ PDF procesado: **{metodo_ocr}**")
                st.info("📋 Revisa y edita los datos:")
                
                df_clean = st.data_editor(df_clean, num_rows="dynamic", use_container_width=True, key="pdf_editor")
                
            else:
                df_raw = pd.read_excel(archivo_a_validar)
                if df_raw.empty:
                    st.error("❌ El archivo está vacío.")
                    st.stop()

                try:
                    df_clean = clean_dataframe(df_raw)
                except ValueError as e:
                    st.error(f"❌ Error: {str(e)}")
                    st.stop()

                if df_clean.empty:
                    st.error("❌ No se encontraron datos válidos.")
                    st.stop()

            # Auto-detection
            norma_detectada, score = detect_norma(df_clean, NORMAS)
            if norma_detectada != norma_seleccionada and score >= 60:
                st.warning(f"⚠️ Detectada: **{norma_detectada}** ({score}%) vs seleccionada: **{norma_seleccionada}**")

            # Validate
            df_resultado = validate_mtc(df_clean, norma_seleccionada, NORMAS)

            n_total    = len(df_resultado)
            n_aprobado = (df_resultado["resultado"] == "APROBADO").sum()
            n_rechazado = (df_resultado["resultado"] == "RECHAZADO").sum()
            n_no_norma = (df_resultado["resultado"] == "NO EN NORMA").sum()
            n_sin_valor = (df_resultado["resultado"] == "SIN VALOR").sum()
            hay_rechazos = n_rechazado > 0

            veredicto = "LOTE RECHAZADO" if hay_rechazos else "LOTE APROBADO"

            if hay_rechazos:
                fallas = df_resultado[df_resultado["resultado"] == "RECHAZADO"]["elemento"].tolist()
                subtexto = f"{n_rechazado} parámetro(s) fuera · {', '.join(fallas)}"
                st.markdown(verdict_banner(veredicto, subtexto, "fail"), unsafe_allow_html=True)
            else:
                subtexto = f"{n_aprobado}/{n_total} parámetros aprobados"
                st.markdown(verdict_banner(veredicto, subtexto, "pass"), unsafe_allow_html=True)

            if n_no_norma > 0:
                elementos_extra = df_resultado[df_resultado["resultado"] == "NO EN NORMA"]["elemento"].tolist()
                st.warning(f"⚠️ {n_no_norma} elemento(s) no en norma: {', '.join(elementos_extra)}")

            if n_sin_valor > 0:
                st.warning(f"⚠️ {n_sin_valor} elemento(s) sin valor")

            k1, k2, k3, k4 = st.columns(4)
            with k1: st.metric("TOTAL", n_total)
            with k2: st.metric("APROBADOS", n_aprobado)
            with k3: st.metric("RECHAZADOS", n_rechazado)
            with k4:
                evaluados = n_aprobado + n_rechazado
                pct = round((n_aprobado / evaluados) * 100) if evaluados > 0 else 0
                st.metric("CUMPLIMIENTO", f"{pct}%")

            st.markdown('<div style="font-size:0.65rem;color:#4a6070;letter-spacing:0.1em;margin:16px 0 8px;">DETALLE POR ELEMENTO</div>', unsafe_allow_html=True)

            def color_resultado(val):
                if val == "APROBADO":     return "background-color: rgba(0,212,160,0.08); color: #00d4a0"
                elif val == "RECHAZADO":  return "background-color: rgba(255,77,106,0.10); color: #ff4d6a"
                elif val == "NO EN NORMA": return "background-color: rgba(245,166,35,0.08); color: #f5a623"
                elif val == "SIN VALOR":  return "background-color: rgba(245,166,35,0.08); color: #f5a623"
                return ""

            df_display = df_resultado[["elemento", "valor", "resultado", "desviacion"]].copy()
            df_display.columns = ["Elemento", "Valor MTC", "Resultado", "Desviación"]
            st.dataframe(
                df_display.style.map(color_resultado, subset=["Resultado"]),
                use_container_width=True, hide_index=True
            )

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            
            try:
                pdf_bytes = generate_pdf(
                    df_resultado, 
                    norma_seleccionada, 
                    veredicto,
                    empresa=st.session_state.name,
                    archivo=archivo_a_validar.name if hasattr(archivo_a_validar, 'name') else "sin_nombre"
                )
                
                col_pdf, col_save = st.columns([3, 1])
                with col_pdf:
                    st.download_button(
                        label="↓ DESCARGAR DICTAMEN PDF",
                        data=pdf_bytes,
                        file_name=f"dictamen_mtc_{norma_seleccionada}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )
                
                with col_save:
                    if st.button("💾 GUARDAR", use_container_width=True, key="save_btn"):
                        try:
                            validacion_id = save_validacion(
                                usuario=st.session_state.name,
                                norma=norma_seleccionada,
                                archivo=archivo_a_validar.name if hasattr(archivo_a_validar, 'name') else "sin_nombre",
                                df_resultado=df_resultado,
                                veredicto=veredicto,
                                metodo=metodo_ingreso,
                                pdf_report=pdf_bytes
                            )
                            if validacion_id:
                                st.success(f"✓ Guardado (ID: {validacion_id}). Ve a 📊 Historial")
                            else:
                                st.error("❌ Error al guardar")
                        except Exception as e:
                            st.error(f"❌ {str(e)}")
                            
            except Exception as e:
                st.error(f"❌ Error PDF: {str(e)}")

            with st.expander("🔧 Debug"):
                st.dataframe(df_resultado, use_container_width=True)

        except Exception as e:
            st.error(f"❌ Error: {str(e)}")
    
    else:
        st.markdown(
            '<div style="background:#0d1218;border:1px solid #1e2835;border-radius:6px;'
            'padding:20px;text-align:center;color:#4a6070;font-size:0.8rem;">'
            '&gt; Sube un archivo para iniciar_'
            '</div>',
            unsafe_allow_html=True
        )

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2: HISTORIAL
# ═══════════════════════════════════════════════════════════════════════════════
with tab_historial:
    st.markdown("""
    <div style="margin-bottom:1.5rem;">
      <div style="font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:700;color:#c8d8e8;">
        Historial de Validaciones
      </div>
      <div style="font-size:0.7rem;color:#4a6070;margin-top:2px;">
        Tu actividad de validaciones MTC
      </div>
    </div>
    """, unsafe_allow_html=True)

    stats = get_stats(st.session_state.name)
    if stats:
        st.markdown('<div style="font-size:0.65rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:8px;">TUS ESTADÍSTICAS</div>', unsafe_allow_html=True)
        s1, s2, s3, s4 = st.columns(4)
        with s1: st.metric("TOTAL", stats.get("total_validaciones", 0))
        with s2: st.metric("APROBADAS", stats.get("total_aprobados", 0))
        with s3: st.metric("RECHAZADAS", stats.get("total_rechazados", 0))
        with s4:
            tasa = stats.get("tasa_aprobacion", 0)
            st.metric("TASA", f"{round(tasa)}%")
        
        st.markdown(f'<div style="font-size:0.65rem;color:#4a6070;margin-top:8px;">Norma más usada: <span style="color:#00d4a0;font-weight:600;">{stats.get("norma_mas_usada", "—")}</span></div>', unsafe_allow_html=True)
        st.markdown("<hr>", unsafe_allow_html=True)

    st.markdown('<div style="font-size:0.65rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:8px;">REGISTROS RECIENTES</div>', unsafe_allow_html=True)
    
    historial = get_historial(st.session_state.name, limit=50)
    
    if historial and len(historial) > 0:
        df_hist = pd.DataFrame(historial).copy()
        
        if "fecha" in df_hist.columns:
            df_hist["fecha"] = pd.to_datetime(df_hist["fecha"]).dt.strftime("%Y-%m-%d %H:%M")
        
        cols_to_show = ["fecha", "norma", "archivo", "n_aprobado", "n_rechazado", "veredicto"]
        df_hist = df_hist[[c for c in cols_to_show if c in df_hist.columns]]
        df_hist.columns = ["Fecha", "Norma", "Archivo", "✓", "✗", "Veredicto"]
        
        def color_veredicto(val):
            if val == "LOTE APROBADO":
                return "background-color: rgba(0,212,160,0.08); color: #00d4a0"
            elif val == "LOTE RECHAZADO":
                return "background-color: rgba(255,77,106,0.10); color: #ff4d6a"
            return ""
        
        st.dataframe(
            df_hist.style.map(color_veredicto, subset=["Veredicto"]),
            use_container_width=True,
            hide_index=True,
            height=300
        )
        
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div style="font-size:0.65rem;color:#ff4d6a;letter-spacing:0.1em;margin-bottom:8px;">PELIGRO</div>', unsafe_allow_html=True)
        
        col_del1, col_del2 = st.columns(2)
        with col_del1:
            if st.button("🗑️ Eliminar registro", use_container_width=True):
                st.info("Ingresa el ID del registro a eliminar")
                delete_id = st.number_input("ID", min_value=1, step=1)
                if st.button("Confirmar", help="No se puede deshacer"):
                    if delete_validacion(delete_id):
                        st.success("✓ Eliminado")
                        st.rerun()
                    else:
                        st.error("❌ Error")
        
        with col_del2:
            if st.button("🗑️ Eliminar TODO", use_container_width=True):
                st.warning("⚠️ Eliminará TODOS tus registros.")
                if st.button("⚠️ CONFIRMAR TOTAL"):
                    if delete_all_usuario(st.session_state.name):
                        st.success("✓ Historial eliminado")
                        st.rerun()
                    else:
                        st.error("❌ Error")
    else:
        st.info("📭 No tienes validaciones registradas.")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3: CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════════════
with tab_config:
    st.markdown("""
    <div style="margin-bottom:1.5rem;">
      <div style="font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:700;color:#c8d8e8;">
        Configuración
      </div>
      <div style="font-size:0.7rem;color:#4a6070;margin-top:2px;">
        Tu cuenta y preferencias
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="font-size:0.65rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:8px;">INFORMACIÓN DE CUENTA</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        st.text_input("Usuario", value=st.session_state.name, disabled=True)
    with col2:
        st.text_input("Email", value=st.session_state.email, disabled=True)
    
    st.markdown("<hr>", unsafe_allow_html=True)
    
    st.markdown('<div style="font-size:0.65rem;color:#4a6070;letter-spacing:0.1em;margin-bottom:8px;">ACERCA DE</div>', unsafe_allow_html=True)
    st.markdown("""
    **MTC Validator v0.3**
    
    Sistema de validación automatizada de certificados metalúrgicos contra normas internacionales.
    
    - 📄 Soporta Excel (.xlsx) y PDFs
    - 🤖 OCR automático para PDFs escaneados con IA (Mistral)
    - 📊 12 normas ASTM/SAE/DIN/API/NMX
    - 💾 Historial persistente con SQLite
    - 📈 Estadísticas personalizadas por usuario
    - 🔐 Sistema de autenticación multi-usuario
    
    **Desarrollado por:** Adrian García · IMT FIME UANL  
    **Contacto:** adaga.tech
    """)
